import openpyxl as opx


def exportar_modelo_para_excel(caminho_arquivo, modelo, df_contas, df_indicadores):
    """
    Exporta os dados tratados e indicadores para um arquivo Excel com placeholders dinâmicos.

    :param caminho_arquivo: Caminho para o arquivo Excel com os modelos.
    :param modelo: Dicionário com as especificações das contas e indicadores a serem exportados.
    :param df_contas: DataFrame com os dados das contas filtradas.
    :param df_indicadores: DataFrame com os indicadores calculados.
    """
    wb = opx.load_workbook(caminho_arquivo)
    aba_nome = modelo["aba"]
    ws = wb[aba_nome]

    # Exportar as contas para a aba especificada
    linha_inicial = modelo["linhas_iniciais"]["contas"]
    for idx, row in df_contas.iterrows():
        ws[f"A{linha_inicial}"] = row['Mes']
        ws[f"B{linha_inicial}"] = row['Empresa']
        ws[f"C{linha_inicial}"] = row['Código']
        ws[f"D{linha_inicial}"] = row['Nome']
        ws[f"E{linha_inicial}"] = row['Valor']
        linha_inicial += 1

    # Exportar indicadores para outra área ou aba, conforme o modelo
    linha_inicial_indicadores = modelo["linhas_iniciais"]["indicadores"]
    for idx, row in df_indicadores.iterrows():
        ws[f"F{linha_inicial_indicadores}"] = row['Indicador']
        ws[f"G{linha_inicial_indicadores}"] = row['Valor']
        linha_inicial_indicadores += 1

    wb.save(caminho_arquivo)
    wb.close()
